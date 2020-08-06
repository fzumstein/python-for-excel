"""This module offers a read and write function to get
2-dimensional lists in and out of Excel files.
"""
import itertools
import datetime as dt

# Optional dependencies
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import pyxlsb
except ImportError:
    pyxlsb = None
try:
    import xlrd
    from xlrd.biffh import error_text_from_code
except ImportError:
    xlrd = None
try:
    import xlwt
    from xlwt.Utils import cell_to_rowcol2
except ImportError:
    xlwt = None
try:
    import xlsxwriter
    from xlsxwriter.utility import xl_cell_to_rowcol
except ImportError:
    xlsxwriter = None


def read(sheet, first_cell='A1', last_cell=None):
    """Read a 2-dimensional list from an Excel range.

    Parameters
    ----------
    sheet : object
        An xlrd, openpyxl or pyxlsb sheet object
    first_cell : str or tuple, optional
        Top-left corner of the Excel range you want to read.
        Can be a string like 'A1' or a row/col tuple like (1, 1),
        default is 'A1'.
    last_cell : str or tuple, optional
        Bottom-right corner of the Excel range you want to read.
        Can be a string like 'A1' or a row/col tuple like (1, 1),
        default is the bottom-right cell of the used range.

    Returns
    -------
    list
        A 2-dimensional list with the values of the Excel range
    """
    # xlrd
    if xlrd and isinstance(sheet, xlrd.sheet.Sheet):
        if last_cell is None:
            # actual range with data, not used range
            last_cell = (sheet.nrows, sheet.ncols)
        if not isinstance(first_cell, tuple):
            first_cell = xlwt.Utils.cell_to_rowcol2(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if not isinstance(last_cell, tuple):
            last_cell = xlwt.Utils.cell_to_rowcol2(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        values = []
        for r in range(first_cell[0] - 1, last_cell[0]):
            row = []
            for c in range(first_cell[1] - 1, last_cell[1]):
                if sheet.cell(r, c).ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        sheet.cell(r, c).value, sheet.book.datemode)
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_EMPTY:
                    value = None
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_ERROR:
                    value = error_text_from_code[sheet.cell(r, c).value]
                else:
                    value = sheet.cell(r, c).value
                row.append(value)
            values.append(row)
        return values
    # OpenPyXL
    elif openpyxl and isinstance(
            sheet,
            (openpyxl.worksheet.worksheet.Worksheet,
             openpyxl.worksheet._read_only.ReadOnlyWorksheet)):
        if last_cell is None:
            # used range
            last_cell = (sheet.max_row, sheet.max_column)
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.cell.coordinate_to_tuple(first_cell)
        if not isinstance(last_cell, tuple):
            last_cell = openpyxl.utils.cell.coordinate_to_tuple(last_cell)
        data = []
        for row in sheet.iter_rows(min_row=first_cell[0], min_col=first_cell[1],
                                   max_row=last_cell[0], max_col=last_cell[1],
                                   values_only=True):
            data.append(list(row))
        return data
    # pyxlsb
    elif pyxlsb and isinstance(sheet, pyxlsb.worksheet.Worksheet):
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.coordinate_to_tuple(first_cell)
        if last_cell and not isinstance(last_cell, tuple):
            last_cell = openpyxl.utils.coordinate_to_tuple(last_cell)
        data = []
        # sheet.rows() is a generator that requires islice to slice it
        for row in itertools.islice(sheet.rows(),
                                    first_cell[0] - 1,
                                    last_cell[0] if last_cell else None):
            data.append([cell.v for cell in row]
                        [first_cell[1] - 1:last_cell[1] if last_cell else None])
        return data
    else:
        raise Exception("Couldn't handle this sheet object!")


def write(sheet, values, first_cell='A1', date_format=None):
    """Write a 2-dimensional list to an Excel range.

    Parameters
    ----------
    sheet : object
        An xlwt, openpyxl or xlsxwriter sheet object
    first_cell : str or tuple, optional
        Top-left corner of the Excel range where you want to write out
        the DataFrame. Can be a string like 'A1' or a row/col tuple
        like (1, 1), default is 'A1'.
    date_format : str, optional
        Only accepted if sheet is an xlwt sheet. By default,
        formats dates like so: 'MM/DD/YY'

    Returns
    -------
    list
        A 2-dimensional list with the values of the Excel range
    """
    # OpenPyXL
    if openpyxl and isinstance(
            sheet, (openpyxl.worksheet.worksheet.Worksheet,
                    openpyxl.worksheet._write_only.WriteOnlyWorksheet)):
        assert date_format is None
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.coordinate_to_tuple(first_cell)
        for i, row in enumerate(values):
            for j, value in enumerate(row):
                sheet.cell(row=first_cell[0] + i,
                           column=first_cell[1] + j).value = value
    # XlsxWriter
    elif xlsxwriter and isinstance(sheet, xlsxwriter.worksheet.Worksheet):
        assert date_format is None
        if isinstance(first_cell, tuple):
            first_cell = first_cell[0] - 1, first_cell[1] - 1
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for r, row_data in enumerate(values):
            sheet.write_row(first_cell[0] + r, first_cell[1], row_data)
    # xlwt
    elif xlwt and isinstance(sheet, xlwt.Worksheet):
        if date_format is None:
            date_format = 'MM/DD/YY'
        date_format = xlwt.easyxf(num_format_str=date_format)
        if isinstance(first_cell, tuple):
            first_cell = (first_cell[0] - 1, first_cell[1] - 1)
        else:
            first_cell = xlwt.Utils.cell_to_rowcol2(first_cell)
        for i, row in enumerate(values):
            for j, cell in enumerate(row):
                if isinstance(cell, (dt.datetime, dt.date)):
                    sheet.write(i + first_cell[0],
                                j + first_cell[1],
                                cell, date_format)
                else:
                    sheet.write(i + first_cell[0],
                                j + first_cell[1], cell)
    else:
        raise Exception("Couldn't handle this sheet object!")
