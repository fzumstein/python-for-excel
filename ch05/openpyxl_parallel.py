import multiprocessing
from itertools import repeat

import openpyxl
import excel


def _read_sheet(filename, sheetname):
    book = openpyxl.load_workbook(filename,
                                  read_only=True, data_only=True)
    sheet = book[sheetname]
    data = excel.read(sheet)
    book.close()
    return sheet.title, data

def load_workbook(filename, sheetnames=None):
    if sheetnames is None:
        book = openpyxl.load_workbook(filename,
                                      read_only=True, data_only=True)
        sheetnames = book.sheetnames
        book.close()
    with multiprocessing.Pool() as pool:
        data = pool.starmap(_read_sheet, zip(repeat(filename), sheetnames))
    return {i[0]: i[1] for i in data}
