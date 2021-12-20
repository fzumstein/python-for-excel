"""This module offers a read and write function to get
2-dimensional lists in and out of Excel files.
"""
import re
import itertools
import datetime as dt

# Optional dependencies
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import pyxlsb
except ImportError:
    pyxlsb = None
try:
    import xlrd
    from xlrd.biffh import error_text_from_code
except ImportError:
    xlrd = None
try:
    import xlwt
except ImportError:
    xlwt = None
try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None


def read(sheet, first_cell="A1", last_cell=None):
    """Read a 2-dimensional list from an Excel range.

    Parameters
    ----------
    sheet : object
        An xlrd, openpyxl or pyxlsb sheet object
    first_cell : str or tuple, optional
        Top-left corner of the Excel range you want to read.
        Can be a string like "A1" or a row/col tuple like (1, 1),
        default is "A1".
    last_cell : str or tuple, optional
        Bottom-right corner of the Excel range you want to read.
        Can be a string like "A1" or a row/col tuple like (1, 1),
        default is the bottom-right cell of the used range.

    Returns
    -------
    list
        A 2-dimensional list with the values of the Excel range
    """
    # xlrd
    if xlrd and isinstance(sheet, xlrd.sheet.Sheet):
        # isinstance returns True if sheet is of type xlrd.sheet.Sheet
        if last_cell is None:
            # actual range with data, not used range
            last_cell = (sheet.nrows, sheet.ncols)
        # Transform "A1" notation into tuples of 1-based indices
        if not isinstance(first_cell, tuple):
            first_cell = xl_cell_to_rowcol(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if not isinstance(last_cell, tuple):
            last_cell = xl_cell_to_rowcol(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        values = []
        for r in range(first_cell[0] - 1, last_cell[0]):
            row = []
            for c in range(first_cell[1] - 1, last_cell[1]):
                # Handle the different cell types
                if sheet.cell(r, c).ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(
                        sheet.cell(r, c).value, sheet.book.datemode)
                elif sheet.cell(r, c).ctype in [xlrd.XL_CELL_EMPTY,
                                                xlrd.XL_CELL_BLANK]:
                    value = None
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_ERROR:
                    value = error_text_from_code[sheet.cell(r, c).value]
                elif sheet.cell(r, c).ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(sheet.cell(r, c).value)
                else:
                    value = sheet.cell(r, c).value
                row.append(value)
            values.append(row)
        return values

    # OpenPyXL
    elif openpyxl and isinstance(
            sheet,
            (openpyxl.worksheet.worksheet.Worksheet,
             openpyxl.worksheet._read_only.ReadOnlyWorksheet)):
        if last_cell is None:
            # used range
            last_cell = (sheet.max_row, sheet.max_column)
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.cell.coordinate_to_tuple(first_cell)
        if not isinstance(last_cell, tuple):
            last_cell = openpyxl.utils.cell.coordinate_to_tuple(last_cell)
        data = []
        for row in sheet.iter_rows(min_row=first_cell[0], min_col=first_cell[1],
                                   max_row=last_cell[0], max_col=last_cell[1],
                                   values_only=True):
            data.append(list(row))
        return data

    # pyxlsb
    elif pyxlsb and isinstance(sheet, pyxlsb.worksheet.Worksheet):
        errors = {"0x0": "#NULL!", "0x7": "#DIV/0!", "0xf": "#VALUE!",
                  "0x17": "#REF!", "0x1d": "#NAME?", "0x24": "#NUM!",
                  "0x2a": "#N/A"}
        if not isinstance(first_cell, tuple):
            first_cell = xl_cell_to_rowcol(first_cell)
            first_cell = (first_cell[0] + 1, first_cell[1] + 1)
        if last_cell and not isinstance(last_cell, tuple):
            last_cell = xl_cell_to_rowcol(last_cell)
            last_cell = (last_cell[0] + 1, last_cell[1] + 1)
        data = []
        # sheet.rows() is a generator that requires islice to slice it
        for row in itertools.islice(sheet.rows(),
                                    first_cell[0] - 1,
                                    last_cell[0] if last_cell else None):
            data.append([errors.get(cell.v, cell.v) for cell in row]
                        [first_cell[1] - 1 : last_cell[1] if last_cell else None])
        return data
    else:
        raise TypeError(f"Couldn't handle sheet of type {type(sheet)}")


def write(sheet, values, first_cell="A1", date_format=None):
    """Write a 2-dimensional list to an Excel range.

    Parameters
    ----------
    sheet : object
        An openpyxl, xlsxwriter or xlwt sheet object. openpyxl's
        write_only=True mode is not supported.
    values : list
        A 2-dimensional list of values
    first_cell : str or tuple, optional
        Top-left corner of the Excel range where you want to write out
        the DataFrame. Can be a string like "A1" or a row/col tuple
        like (1, 1), default is "A1".
    date_format : str, optional
        Only accepted if sheet is an openpyxl or xlwt sheet. By default,
        formats dates in the following format: "mm/dd/yy". For xlsxwriter,
        set the format when you instantiate a Workbook by providing:
        options={"default_date_format": "mm/dd/yy"}
    """
    # OpenPyXL
    if openpyxl and isinstance(
            sheet, openpyxl.worksheet.worksheet.Worksheet):
        if date_format is None:
            date_format = "mm/dd/yy"
        if not isinstance(first_cell, tuple):
            first_cell = openpyxl.utils.coordinate_to_tuple(first_cell)
        for i, row in enumerate(values):
            for j, value in enumerate(row):
                cell = sheet.cell(row=first_cell[0] + i,
                                  column=first_cell[1] + j)
                cell.value = value
                if date_format and isinstance(value, (dt.datetime, dt.date)):
                    cell.number_format = date_format

    # XlsxWriter
    elif xlsxwriter and isinstance(sheet, xlsxwriter.worksheet.Worksheet):
        if date_format is not None:
            raise ValueError("date_format must be set as Workbook option")
        if isinstance(first_cell, tuple):
            first_cell = first_cell[0] - 1, first_cell[1] - 1
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for r, row_data in enumerate(values):
            sheet.write_row(first_cell[0] + r, first_cell[1], row_data)

    # xlwt
    elif xlwt and isinstance(sheet, xlwt.Worksheet):
        if date_format is None:
            date_format = "mm/dd/yy"
        date_format = xlwt.easyxf(num_format_str=date_format)
        if isinstance(first_cell, tuple):
            first_cell = (first_cell[0] - 1, first_cell[1] - 1)
        else:
            first_cell = xl_cell_to_rowcol(first_cell)
        for i, row in enumerate(values):
            for j, cell in enumerate(row):
                if isinstance(cell, (dt.datetime, dt.date)):
                    sheet.write(i + first_cell[0], j + first_cell[1],
                                cell, date_format)
                else:
                    sheet.write(i + first_cell[0], j + first_cell[1],
                                cell)
    else:
        raise TypeError(f"Couldn't handle sheet of type {type(sheet)}")


def xl_cell_to_rowcol(cell_str):
    """
    Convert a cell reference in A1 notation to a zero indexed row and column.

    Args:
       cell_str:  A1 style string.

    Returns:
        row, col: Zero indexed cell row and column indices.

    This function is from XlsxWriter
    Copyright (c) 2013-2020, John McNamara <jmcnamara@cpan.org>
    All rights reserved.

    Redistribution and use in source and binary forms, with or without
    modification, are permitted provided that the following conditions are met:

    1. Redistributions of source code must retain the above copyright notice, this
       list of conditions and the following disclaimer.
    2. Redistributions in binary form must reproduce the above copyright notice,
       this list of conditions and the following disclaimer in the documentation
       and/or other materials provided with the distribution.

    THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND
    ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED
    WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE
    DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT OWNER OR CONTRIBUTORS BE LIABLE FOR
    ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES
    (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES;
    LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND
    ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT
    (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS
    SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.

    The views and conclusions contained in the software and documentation are those
    of the authors and should not be interpreted as representing official policies,
    either expressed or implied, of the FreeBSD Project.

    """
    if not cell_str:
        return 0, 0

    match = re.compile(r"(\$?)([A-Z]{1,3})(\$?)(\d+)").match(cell_str)
    col_str = match.group(2)
    row_str = match.group(4)

    # Convert base26 column string to number.
    expn = 0
    col = 0
    for char in reversed(col_str):
        col += (ord(char) - ord("A") + 1) * (26 ** expn)
        expn += 1

    # Convert 1-index to zero-index
    row = int(row_str) - 1
    col -= 1

    return row, col
