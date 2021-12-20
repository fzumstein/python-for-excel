import multiprocessing
from itertools import repeat

import openpyxl
import excel


def _read_sheet(filename, sheetname):
    # The leading underscore in the function name is used by convention
    # to mark it as "private", i.e., it shouldn't be used directly outside
    # of this module.
    book = openpyxl.load_workbook(filename,
                                  read_only=True, data_only=True)
    sheet = book[sheetname]
    data = excel.read(sheet)
    book.close()
    return sheet.title, data

def load_workbook(filename, sheetnames=None):
    if sheetnames is None:
        book = openpyxl.load_workbook(filename,
                                      read_only=True, data_only=True)
        sheetnames = book.sheetnames
        book.close()
    with multiprocessing.Pool() as pool:
        # By default, Pool spawns as many processes as there are CPU cores.
        # starmap maps a tuple of arguments to a function. The zip expression
        # produces a list with tuples of the following form:
        # [('filename.xlsx', 'Sheet1'), ('filename.xlsx', 'Sheet2)]
        data = pool.starmap(_read_sheet, zip(repeat(filename), sheetnames))
    return {i[0]: i[1] for i in data}
