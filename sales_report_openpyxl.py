from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties


# Directory of this file
this_dir = Path(__file__).resolve().parent

# Read in all files
parts = []
for path in (this_dir / "sales_data").rglob("*.xls*"):
    print(f'Reading {path.name}')
    part = pd.read_excel(path)
    parts.append(part)

# Combine the DataFrames from each file into a single DataFrame
df = pd.concat(parts)

# Pivot each store into a column and sum up all transactions per date
pivot = pd.pivot_table(df,
                       index="transaction_date", columns="store",
                       values="amount", aggfunc="sum")

# Resample to end of month and assign an index name
summary = pivot.resample("M").sum()
summary.index.name = "Month"

# Sort columns by total revenue
summary = summary.loc[:, summary.sum().sort_values().index]

# Add row and column totals: Using "append" together with "rename"
# is a convenient way to add a row to the bottom of a DataFrame
summary.loc[:, "Total"] = summary.sum(axis=1)
summary = summary.append(summary.sum(axis=0).rename("Total"))

#### Write summary report to Excel file ####

# DataFrame position and number of rows/columns
# openpxyl uses 1-based indices
startrow, startcol = 3, 2
nrows, ncols = summary.shape

# Starting with pandas 1.3.0, the following line will raise a FutureWarning.
# To fix this, replace write_only=True with engine_kwargs={"write_only": True}
with pd.ExcelWriter(this_dir / "sales_report_openpyxl.xlsx",
                    engine="openpyxl", write_only=True) as writer:
    # pandas uses 0-based indices
    summary.to_excel(writer, sheet_name="Sheet1",
                     startrow=startrow - 1, startcol=startcol - 1)

    # Get openpyxl book and sheet object
    book = writer.book
    sheet = writer.sheets["Sheet1"]

    # Set title
    sheet.cell(row=1, column=startcol, value="Sales Report")
    sheet.cell(row=1, column=startcol).font = Font(size=24, bold=True)

    # Sheet formatting
    sheet.sheet_view.showGridLines = False

    # Format the DataFrame with
    # - number format
    # - column width
    # - conditional formatting
    for row in range(startrow + 1, startrow + nrows + 1):
        for col in range(startcol + 1, startcol + ncols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="center")

    for cell in sheet["B"]:
        cell.number_format = "mmm yy"

    for col in range(startcol, startcol + ncols + 1):
        cell = sheet.cell(row=startrow, column=col)
        sheet.column_dimensions[cell.column_letter].width = 14

    first_cell = sheet.cell(row=startrow + 1, column=startcol + 1)
    last_cell = sheet.cell(row=startrow + nrows, column=startcol + ncols)
    range_address = f"{first_cell.coordinate}:{last_cell.coordinate}"
    sheet.conditional_formatting.add(range_address,
                                     CellIsRule(operator="lessThan",
                                                formula=["20000"],
                                                stopIfTrue=True,
                                                font=Font(color="E93423")))

    # Chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Sales per Month and Store"
    chart.height = 11.5
    chart.width = 20.5

    # Add each column as a series, ignoring total row and col
    data = Reference(sheet, min_col=startcol + 1, min_row=startrow,
                     max_row=startrow + nrows - 1,
                     max_col=startcol + ncols - 1)
    categories = Reference(sheet, min_col=startcol, min_row=startrow + 1,
                           max_row=startrow + nrows - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    cell = sheet.cell(row=startrow + nrows + 2, column=startcol)
    sheet.add_chart(chart=chart, anchor=cell.coordinate)

    # Chart formatting
    chart.y_axis.title = "Sales"
    chart.x_axis.title = summary.index.name
    # Hide y-axis line: spPR stands for ShapeProperties 
    chart.y_axis.spPr = GraphicalProperties(ln=LineProperties(noFill=True))
